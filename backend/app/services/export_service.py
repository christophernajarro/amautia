"""Export exam results to PDF and Excel."""
import io
from datetime import datetime


def export_results_excel(exam_title: str, results: list[dict], stats: dict) -> bytes:
    """Generate Excel workbook with exam results."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    # Colors
    indigo = "4F46E5"
    green = "10B981"
    red = "EF4444"
    amber = "F59E0B"
    light_blue = "EEF2FF"

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Resultados: {exam_title}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=indigo)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Stats row
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Promedio: {stats.get('average', 0):.1f}% | Total: {stats.get('total_students', 0)} | Corregidos: {stats.get('corrected', 0)} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].fill = PatternFill("solid", fgColor=light_blue)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Headers
    headers = ["#", "Alumno", "Email", "Puntaje", "Porcentaje", "Estado", "Retroalimentación"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="334155")
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 22

    # Data rows
    for i, r in enumerate(results, 1):
        row = 3 + i
        pct = r.get("percentage") or 0
        color = green if pct >= 70 else amber if pct >= 50 else red

        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=r.get("student_name", "Sin nombre"))
        ws.cell(row=row, column=3, value=r.get("student_email", ""))
        score_cell = ws.cell(row=row, column=4, value=r.get("total_score"))
        score_cell.alignment = Alignment(horizontal="center")
        pct_cell = ws.cell(row=row, column=5, value=f"{pct:.1f}%" if pct else "—")
        pct_cell.font = Font(bold=True, color=color)
        pct_cell.alignment = Alignment(horizontal="center")
        status = r.get("status", "")
        ws.cell(row=row, column=6, value="✓ Corregido" if status == "corrected" else status).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=7, value=r.get("feedback", "")[:200] if r.get("feedback") else "")

        if i % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F8FAFC")

    # Column widths
    from openpyxl.utils import get_column_letter
    col_widths = [5, 25, 30, 10, 12, 15, 50]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Stats sheet
    ws2 = wb.create_sheet("Estadísticas")
    ws2["A1"] = "Estadística"
    ws2["B1"] = "Valor"
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    stat_rows = [
        ("Total alumnos", stats.get("total_students", 0)),
        ("Corregidos", stats.get("corrected", 0)),
        ("Promedio", f"{stats.get('average', 0):.1f}%"),
        ("Nota máxima", f"{stats.get('max_score', 0):.1f}%"),
        ("Nota mínima", f"{stats.get('min_score', 0):.1f}%"),
    ]
    for i, (k, v) in enumerate(stat_rows, 2):
        ws2[f"A{i}"] = k
        ws2[f"B{i}"] = v
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_results_pdf(exam_title: str, results: list[dict], stats: dict) -> bytes:
    """Generate PDF report with exam results."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm,
                            leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=18,
                                 textColor=colors.HexColor("#4F46E5"), spaceAfter=4)
    elements.append(Paragraph(f"Resultados: {exam_title}", title_style))

    # Stats summary
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10, textColor=colors.grey)
    avg = stats.get("average", 0)
    elements.append(Paragraph(
        f"Promedio: {avg:.1f}% · Total: {stats.get('total_students', 0)} alumnos · "
        f"Corregidos: {stats.get('corrected', 0)} · Nota máx: {stats.get('max_score', 0):.1f}% · "
        f"Nota mín: {stats.get('min_score', 0):.1f}% · "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        sub_style))
    elements.append(Spacer(1, 8*mm))

    # Table
    headers = ["#", "Alumno", "Puntaje", "%", "Estado"]
    data = [headers]
    for i, r in enumerate(results, 1):
        pct = r.get("percentage") or 0
        data.append([
            str(i),
            r.get("student_name", "—")[:30],
            str(r.get("total_score", "—")),
            f"{pct:.1f}%" if pct else "—",
            "✓ Corregido" if r.get("status") == "corrected" else r.get("status", ""),
        ])

    table = Table(data, colWidths=[12*mm, 70*mm, 22*mm, 22*mm, 30*mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (1, 1), (1, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    return buf.read()
